import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import os
from datetime import datetime
from decimal import Decimal
from ttkthemes import ThemedTk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sys

class SmoothScrollbar(ttk.Scrollbar):
    """自定义平滑滚动条"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._smooth_scroll = 0.1
        self._last_y = 0
        self._last_time = 0
        
    def set(self, lo, hi):
        if float(lo) <= 0.0 and float(hi) >= 1.0:
            self.grid_remove()
        else:
            self.grid()
        super().set(lo, hi)

class NFeAnalyzer(ThemedTk):
    def __init__(self):
        super().__init__(theme="cosmo")
        
        self.title("NFe 数据分析器")
        self.geometry("1800x900")
        
        # 初始化数据存储
        self.loaded_files = {}
        self.invoice_data = {}
        self.summary_labels = {}
        self.summary_data = {
            "total_invoices": 0,
            "total_products": 0,
            "total_amount": Decimal('0'),
            "total_icms": Decimal('0'),
            "total_pis": Decimal('0'),
            "total_cofins": Decimal('0'),
            "total_tax": Decimal('0'),
            "avg_product_price": Decimal('0'),
            "max_product_price": Decimal('0'),
            "min_product_price": Decimal('0'),
            "total_discount": Decimal('0'),
            "unique_customers": set(),
            "unique_products": set(),
            "tax_percentage": Decimal('0')
        }
        
        # 设置防抖动
        self._scroll_update_pending = False
        self._last_scroll_time = 0
        self._scroll_delay = 10
        
        # 设置样式
        self.style = ttk.Style()
        
        # 主题颜色
        self.colors = {
            'primary': '#2196f3',
            'secondary': '#757575',
            'success': '#4caf50',
            'info': '#00bcd4',
            'warning': '#ff9800',
            'danger': '#f44336',
            'light': '#f5f5f5',
            'dark': '#212121',
            'border': '#e0e0e0',
            'hover': '#1976d2'
        }
        
        # 配置基本样式
        self.style.configure(
            "Card.TLabelframe",
            borderwidth=0,
            relief="flat",
            background=self.colors['light']
        )
        
        self.style.configure(
            "Card.TLabelframe.Label",
            font=("Arial", 12, "bold"),
            foreground=self.colors['dark'],
            background=self.colors['light'],
            padding=(10, 5)
        )
        
        # 配置标签样式
        self.style.configure(
            "Title.TLabel",
            font=("Arial", 12, "bold"),
            foreground=self.colors['dark'],
            background=self.colors['light'],
            padding=(5, 5)
        )
        
        self.style.configure(
            "Value.TLabel",
            font=("Arial", 11),
            foreground=self.colors['dark'],
            background=self.colors['light'],
            padding=(5, 2)
        )
        
        self.style.configure(
            "Header.TLabel",
            font=("Arial", 10, "bold"),
            foreground=self.colors['dark'],
            background=self.colors['light'],
            padding=(5, 2)
        )
        
        # 配置按钮样式
        for style_name, color in [
            ('primary', 'primary'),
            ('secondary', 'secondary'),
            ('success', 'success'),
            ('info', 'info'),
            ('warning', 'warning'),
            ('danger', 'danger')
        ]:
            self.style.configure(
                f"{style_name}.TButton",
                font=("Arial", 10),
                background=self.colors[color],
                foreground="white",
                padding=(15, 8),
                borderwidth=0,
                relief="flat"
            )
            
            # 添加鼠标悬停效果
            self.style.map(
                f"{style_name}.TButton",
                background=[('active', self.colors['hover'])],
                relief=[('pressed', 'sunken')]
            )
        
        # 配置树形视图样式
        self.style.configure(
            "Treeview",
            background=self.colors['light'],
            foreground=self.colors['dark'],
            fieldbackground=self.colors['light'],
            borderwidth=0,
            relief="flat",
            rowheight=30
        )
        
        self.style.configure(
            "Treeview.Heading",
            font=("Arial", 10, "bold"),
            background=self.colors['primary'],
            foreground="white",
            padding=(10, 5),
            borderwidth=0
        )
        
        self.style.map(
            "Treeview",
            background=[('selected', self.colors['primary'])],
            foreground=[('selected', 'white')]
        )
        
        # 配置滚动条样式
        self.style.configure(
            "Smooth.Vertical.TScrollbar",
            background=self.colors['primary'],
            troughcolor=self.colors['light'],
            borderwidth=0,
            relief="flat",
            width=10,
            arrowsize=13
        )
        
        self.style.configure(
            "Smooth.Horizontal.TScrollbar",
            background=self.colors['primary'],
            troughcolor=self.colors['light'],
            borderwidth=0,
            relief="flat",
            width=10,
            arrowsize=13
        )
        
        # 创建主框架
        self.main_frame = ttk.Frame(self, style="Card.TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 创建控制面板
        self.create_control_panel()
        
        # 创建Notebook
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # 创建各个标签页
        self.create_data_page()
        self.create_invoice_analysis_page()
        self.create_summary_page()
        
        # 绑定标签页切换事件
        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)
        
        # 初始化数据
        self._init_data()
        
    def _init_data(self):
        """重置数据到初始状态"""
        # 清除数据
        self.loaded_files.clear()
        self.invoice_data.clear()
        
        # 重置防抖动设置
        self._scroll_update_pending = False
        self._last_scroll_time = 0
        
    def _on_tab_changed(self, event):
        """处理标签页切换事件"""
        tab = self.notebook.select()
        tab_text = self.notebook.tab(tab, "text")
        
        # 添加切换动画效果
        self.notebook.update_idletasks()
        
        # 更新当前页面数据
        if tab_text == "统计分析":
            self.update_summary_display()
        elif tab_text == "发票分析":
            self.update_invoice_display()
        
    def create_control_panel(self):
        """创建控制面板"""
        control_frame = ttk.Frame(self.main_frame, style="Card.TFrame")
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 文件操作按钮
        btn_frame = ttk.Frame(control_frame, style="Card.TFrame")
        btn_frame.pack(side=tk.LEFT, padx=10, pady=5)
        
        buttons = [
            ("选择NFe文件", self.select_files, "primary"),
            ("导出Excel", self.export_to_excel, "info"),
            ("清除所有", self.clear_all, "secondary")
        ]
        
        for text, command, style in buttons:
            btn = ttk.Button(
            btn_frame,
                text=text,
                command=command,
                style=f"{style}.TButton",
            width=15
        )
            btn.pack(side=tk.LEFT, padx=5)
            
            # 添加鼠标悬停效果
            btn.bind('<Enter>', lambda e, b=btn: self._on_button_hover(e, b))
            btn.bind('<Leave>', lambda e, b=btn: self._on_button_leave(e, b))
        
        # 文件列表
        file_list_frame = ttk.LabelFrame(
            control_frame,
            text="已加载文件",
            style="Card.TLabelframe"
        )
        file_list_frame.pack(fill=tk.X, padx=(20, 10), pady=5)
        
        # 创建文件树形视图
        self.file_tree = ttk.Treeview(
            file_list_frame,
            columns=("文件名", "状态"),
            show="headings",
            height=3,
            style="Treeview"
        )
        
        # 设置列
        self.file_tree.heading("文件名", text="文件名")
        self.file_tree.heading("状态", text="状态")
        self.file_tree.column("文件名", width=300)
        self.file_tree.column("状态", width=100)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(
            file_list_frame,
            orient=tk.VERTICAL,
            command=self.file_tree.yview,
            style="Smooth.Vertical.TScrollbar"
        )
        
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        self.file_tree.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        
        # 绑定鼠标滚轮事件
        self.file_tree.bind("<Enter>", self._bind_mousewheel)
        self.file_tree.bind("<Leave>", self._unbind_mousewheel)
        
    def _on_button_hover(self, event, button):
        """处理按钮悬停事件"""
        style = str(button.cget('style')).split('.')[0]
        button.configure(cursor="hand2")
        self.style.configure(
            f"{style}.TButton",
            background=self.colors['hover']
        )
        
    def _on_button_leave(self, event, button):
        """处理按钮离开事件"""
        style = str(button.cget('style')).split('.')[0]
        button.configure(cursor="")
        self.style.configure(
            f"{style}.TButton",
            background=self.colors[style.lower()]
        )
        
    def create_data_page(self):
        """创建数据明细页面"""
        data_page = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(data_page, text="数据明细")
        
        # 创建数据显示区域
        data_frame = ttk.Frame(data_page, style="Card.TFrame")
        data_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建容器框架
        container = ttk.Frame(data_frame, style="Card.TFrame")
        container.pack(fill=tk.BOTH, expand=True)
        
        # 定义列
        columns = (
            "发票号码", "系列号", "发票日期", "操作性质", "打印类型", "出具类型",
            "客户名称", "客户CNPJ", "客户地址",
            "产品代码", "产品名称", "NCM", "数量", "单价", "总价",
            "ICMS原始税率", "ICMS应纳税额",
            "PIS税率", "PIS应纳税额",
            "COFINS税率", "COFINS应纳税额"
        )
        
        # 创建树形视图
        self.tree = ttk.Treeview(
            container,
            columns=columns,
            show="headings",
            style="Treeview"
        )
        
        # 设置列宽和标题
        widths = {
            "发票号码": 80,
            "系列号": 60,
            "发票日期": 100,
            "操作性质": 150,
            "打印类型": 80,
            "出具类型": 80,
            "客户名称": 200,
            "客户CNPJ": 120,
            "客户地址": 250,
            "产品代码": 100,
            "产品名称": 250,
            "NCM": 80,
            "数量": 80,
            "单价": 100,
            "总价": 100,
            "ICMS原始税率": 100,
            "ICMS应纳税额": 100,
            "PIS税率": 80,
            "PIS应纳税额": 100,
            "COFINS税率": 80,
            "COFINS应纳税额": 100
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=widths.get(col, 100))
        
        # 添加滚动条
        v_scrollbar = ttk.Scrollbar(
            container,
            orient=tk.VERTICAL,
            command=self.tree.yview,
            style="Smooth.Vertical.TScrollbar"
        )
        
        h_scrollbar = ttk.Scrollbar(
            container,
            orient=tk.HORIZONTAL,
            command=self.tree.xview,
            style="Smooth.Horizontal.TScrollbar"
        )
        
        # 配置滚动绑定
        self.tree.configure(
            yscrollcommand=lambda *args: self._debounced_scroll(v_scrollbar, *args),
            xscrollcommand=lambda *args: self._debounced_scroll(h_scrollbar, *args)
        )
        
        # 使用网格布局
        self.tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 配置网格权重
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        # 绑定事件
        self.tree.bind("<Double-1>", self._on_tree_double_click)
        
        # 绑定滚动事件
        self.tree.bind("<Enter>", lambda e: (self._bind_mousewheel(e), self._bind_shift_mousewheel(e)))
        self.tree.bind("<Leave>", lambda e: (self._unbind_mousewheel(e), self._unbind_shift_mousewheel(e)))
        
    def _on_tree_double_click(self, event):
        """处理树形视图双击事件"""
        item = self.tree.selection()[0]
        values = self.tree.item(item)["values"]
        if values:
            # 显示详细信息对话框
            self._show_detail_dialog(values)
            
    def _show_detail_dialog(self, values):
        """显示详细信息对话框"""
        dialog = tk.Toplevel(self)
        dialog.title("详细信息")
        dialog.geometry("600x400")
        
        # 设置对话框样式
        frame = ttk.Frame(dialog, style="Card.TFrame")
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 显示详细信息
        headers = [
            "发票号码", "系列号", "发票日期", "操作性质", "打印类型", "出具类型",
            "客户名称", "客户CNPJ", "客户地址",
            "产品代码", "产品名称", "NCM", "数量", "单价", "总价",
            "ICMS原始税率", "ICMS应纳税额",
            "PIS税率", "PIS应纳税额",
            "COFINS税率", "COFINS应纳税额"
        ]
        
        for i, (header, value) in enumerate(zip(headers, values)):
            row = ttk.Frame(frame, style="Card.TFrame")
            row.pack(fill=tk.X, pady=2)
            
            ttk.Label(
                row,
                text=f"{header}:",
                style="Header.TLabel"
            ).pack(side=tk.LEFT, padx=5)
            
            ttk.Label(
                row,
                text=str(value),
                style="Value.TLabel"
            ).pack(side=tk.RIGHT, padx=5)
        
        # 添加关闭按钮
        ttk.Button(
            frame,
            text="关闭",
            command=dialog.destroy,
            style="primary.TButton"
        ).pack(pady=20)
        
        # 使对话框居中显示
        dialog.transient(self)
        dialog.grab_set()
        self.wait_window(dialog)
        
    def create_invoice_analysis_page(self):
        """创建发票分析页面"""
        invoice_page = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(invoice_page, text="发票分析")
        
        # 创建发票列表框架
        invoice_frame = ttk.LabelFrame(
            invoice_page,
            text="发票列表",
            style="Card.TLabelframe"
        )
        invoice_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建工具栏
        toolbar = ttk.Frame(invoice_frame, style="Card.TFrame")
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        # 添加搜索框
        search_frame = ttk.Frame(toolbar, style="Card.TFrame")
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        ttk.Label(
            search_frame,
            text="搜索:",
            style="Header.TLabel"
        ).pack(side=tk.LEFT, padx=5)
        
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(
            search_frame,
            textvariable=self.search_var,
            width=40
        )
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # 添加过滤按钮
        filter_btn = ttk.Button(
            toolbar,
            text="过滤",
            style="info.TButton",
            width=10,
            command=self._apply_filter
        )
        filter_btn.pack(side=tk.RIGHT, padx=5)
        
        # 创建表格容器
        table_container = ttk.Frame(invoice_frame, style="Card.TFrame")
        table_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建发票树形视图
        columns = (
            "发票号码", "发票日期", "客户名称", "客户CNPJ", "客户地址", 
            "产品总价", "运费", "折扣",
            "ICMS税率", "ICMS税额", 
            "PIS税率", "PIS税额", 
            "COFINS税率", "COFINS税额"
        )
        
        self.invoice_tree = ttk.Treeview(
            table_container,
            columns=columns,
            show="headings",
            style="Treeview",
            height=20
        )
        
        # 设置列宽和标题
        widths = {
            "发票号码": 100,
            "发票日期": 100,
            "客户名称": 200,
            "客户CNPJ": 130,
            "客户地址": 300,
            "产品总价": 120,
            "运费": 80,
            "折扣": 80,
            "ICMS税率": 80,
            "ICMS税额": 120,
            "PIS税率": 80,
            "PIS税额": 120,
            "COFINS税率": 80,
            "COFINS税额": 120
        }
        
        for col in columns:
            self.invoice_tree.heading(col, text=col)
            self.invoice_tree.column(col, width=widths.get(col, 100))
        
        # 添加滚动条
        v_scrollbar = ttk.Scrollbar(
            table_container,
            orient=tk.VERTICAL,
            command=self.invoice_tree.yview,
            style="Smooth.Vertical.TScrollbar"
        )
        
        h_scrollbar = ttk.Scrollbar(
            table_container,
            orient=tk.HORIZONTAL,
            command=self.invoice_tree.xview,
            style="Smooth.Horizontal.TScrollbar"
        )
        
        # 配置滚动绑定
        self.invoice_tree.configure(
            yscrollcommand=lambda *args: self._debounced_scroll(v_scrollbar, *args),
            xscrollcommand=lambda *args: self._debounced_scroll(h_scrollbar, *args)
        )
        
        # 使用pack布局
        self.invoice_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 绑定事件
        self.invoice_tree.bind("<Double-1>", self._on_invoice_double_click)
        self.search_var.trace("w", lambda *args: self._on_search_change())
        
        # 绑定滚动事件
        self.invoice_tree.bind("<Enter>", lambda e: (self._bind_mousewheel(e), self._bind_shift_mousewheel(e)))
        self.invoice_tree.bind("<Leave>", lambda e: (self._unbind_mousewheel(e), self._unbind_shift_mousewheel(e)))
        
    def _apply_filter(self):
        """应用过滤器"""
        search_text = self.search_var.get().lower()
        
        # 清除现有项目
        for item in self.invoice_tree.get_children():
            self.invoice_tree.delete(item)
        
        # 重新添加匹配的项目
        for invoice_number, data in self.invoice_data.items():
            if (search_text in invoice_number.lower() or
                search_text in data["customer_name"].lower() or
                search_text in data["customer_cnpj"].lower()):
                
                self.invoice_tree.insert("", tk.END, values=(
                    invoice_number,
                    data["invoice_date"],
                    data["customer_name"],
                    data["customer_cnpj"],
                    data["customer_address"],
                    f"R$ {float(data['total_amount']):,.2f}",
                    f"R$ {float(data['freight']):,.2f}",
                    f"R$ {float(data['discount']):,.2f}",
                    f"{data['icms_rate']}%",
                    f"R$ {float(data['icms_amount']):,.2f}",
                    f"{data['pis_rate']}%",
                    f"R$ {float(data['pis_amount']):,.2f}",
                    f"{data['cofins_rate']}%",
                    f"R$ {float(data['cofins_amount']):,.2f}"
                ))
        
    def _on_search_change(self):
        """处理搜索文本变化"""
        if hasattr(self, "_search_after"):
            self.after_cancel(self._search_after)
        self._search_after = self.after(300, self._apply_filter)
        
    def _on_invoice_double_click(self, event):
        """处理发票双击事件"""
        item = self.invoice_tree.selection()[0]
        values = self.invoice_tree.item(item)["values"]
        if values:
            self._show_invoice_detail(values[0])  # 传递发票号码
            
    def _show_invoice_detail(self, invoice_number):
        """显示发票详细信息"""
        if invoice_number not in self.invoice_data:
            return
            
        data = self.invoice_data[invoice_number]
        
        dialog = tk.Toplevel(self)
        dialog.title(f"发票详情 - {invoice_number}")
        dialog.geometry("800x600")
        
        # 设置对话框样式
        frame = ttk.Frame(dialog, style="Card.TFrame")
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 创建选项卡
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # 基本信息页
        basic_page = ttk.Frame(notebook, style="Card.TFrame")
        notebook.add(basic_page, text="基本信息")
        
        # 显示基本信息
        basic_info = [
            ("发票号码", invoice_number),
            ("发票日期", data["invoice_date"]),
            ("客户名称", data["customer_name"]),
            ("客户CNPJ", data["customer_cnpj"]),
            ("客户地址", data["customer_address"]),
            ("产品总价", f"R$ {float(data['total_amount']):,.2f}"),
            ("运费", f"R$ {float(data['freight']):,.2f}"),
            ("折扣", f"R$ {float(data['discount']):,.2f}")
        ]
        
        for i, (label, value) in enumerate(basic_info):
            row = ttk.Frame(basic_page, style="Card.TFrame")
            row.pack(fill=tk.X, pady=2)
            
            ttk.Label(
                row,
                text=f"{label}:",
                style="Header.TLabel"
            ).pack(side=tk.LEFT, padx=5)
            
            ttk.Label(
                row,
                text=str(value),
                style="Value.TLabel"
            ).pack(side=tk.RIGHT, padx=5)
        
        # 税收信息页
        tax_page = ttk.Frame(notebook, style="Card.TFrame")
        notebook.add(tax_page, text="税收信息")
        
        # 显示税收信息
        tax_info = [
            ("ICMS税率", f"{data['icms_rate']}%"),
            ("ICMS税额", f"R$ {float(data['icms_amount']):,.2f}"),
            ("PIS税率", f"{data['pis_rate']}%"),
            ("PIS税额", f"R$ {float(data['pis_amount']):,.2f}"),
            ("COFINS税率", f"{data['cofins_rate']}%"),
            ("COFINS税额", f"R$ {float(data['cofins_amount']):,.2f}")
        ]
        
        for i, (label, value) in enumerate(tax_info):
            row = ttk.Frame(tax_page, style="Card.TFrame")
            row.pack(fill=tk.X, pady=2)
            
            ttk.Label(
                row,
                text=f"{label}:",
                style="Header.TLabel"
            ).pack(side=tk.LEFT, padx=5)
            
            ttk.Label(
                row,
                text=str(value),
                style="Value.TLabel"
            ).pack(side=tk.RIGHT, padx=5)
        
        # 添加关闭按钮
        ttk.Button(
            frame,
            text="关闭",
            command=dialog.destroy,
            style="primary.TButton"
        ).pack(pady=20)
        
        # 使对话框居中显示
        dialog.transient(self)
        dialog.grab_set()
        self.wait_window(dialog)

    def select_files(self):
        """选择NFe文件"""
        files = filedialog.askopenfilenames(
            title="选择NFe文件",
            filetypes=[("XML files", "*.xml")],
            multiple=True
        )
        
        if not files:
            return
        
        for file_path in files:
            file_name = os.path.basename(file_path)
            try:
                # 检查文件是否已加载
                if file_path in self.loaded_files:
                    messagebox.showwarning("警告", f"文件 {file_name} 已经加载")
                    continue
                
                # 解析XML文件
                tree = ET.parse(file_path)
                
                # 添加到文件列表
                self.file_tree.insert("", tk.END, values=(file_name, "处理中..."))
                self.loaded_files[file_path] = "处理中"
                
                # 处理XML数据
                self.process_xml(tree)
                
                # 更新文件状态
                for item in self.file_tree.get_children():
                    if self.file_tree.item(item)["values"][0] == file_name:
                        self.file_tree.item(item, values=(file_name, "已完成"))
                        self.loaded_files[file_path] = "已完成"
                        break
            except ET.ParseError:
                messagebox.showerror("错误", f"文件 {file_name} 不是有效的XML文件")
                continue
            except Exception as e:
                messagebox.showerror("错误", f"处理文件 {file_name} 时出错: {str(e)}")
                print(f"错误详情: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def clear_all(self):
        """清除所有数据"""
        if messagebox.askyesno("确认", "确定要清除所有数据吗？"):
            # 清除文件列表
            for item in self.file_tree.get_children():
                self.file_tree.delete(item)
            
            # 清除数据表格
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # 清除发票表格
            for item in self.invoice_tree.get_children():
                self.invoice_tree.delete(item)
            
            # 清除数据
            self.loaded_files.clear()
            self.invoice_data.clear()
            self.summary_labels.clear()
            
            # 重置汇总数据
            self.summary_data = {
                "total_invoices": 0,
                "total_products": 0,
                "total_amount": Decimal('0'),
                "total_icms": Decimal('0'),
                "total_pis": Decimal('0'),
                "total_cofins": Decimal('0'),
                "total_tax": Decimal('0'),
                "avg_product_price": Decimal('0'),
                "max_product_price": Decimal('0'),
                "min_product_price": Decimal('0'),
                "total_discount": Decimal('0'),
                "unique_customers": set(),
                "unique_products": set(),
                "tax_percentage": Decimal('0')
            }
            
            # 更新显示
            self.update_summary_display()
            
            # 重置防抖动设置
            self._scroll_update_pending = False
            self._last_scroll_time = 0

    def safe_find_text(self, element, path, ns, default=""):
        """安全获取XML元素文本"""
        if element is None:
            return default
        node = element.find(path, ns)
        return node.text if node is not None else default
    
    def safe_find_decimal(self, element, path, ns):
        """安全获取XML元素的Decimal值"""
        if element is None:
            return Decimal('0')
        node = element.find(path, ns)
        try:
            return Decimal(node.text) if node is not None else Decimal('0')
        except:
            return Decimal('0')
    
    def format_date(self, date_str):
        """格式化日期字符串"""
        try:
            if not date_str:
                return ""
            # 处理带时区的日期时间
            if "T" in date_str:
                dt = datetime.strptime(date_str.split("T")[0], "%Y-%m-%d")
            else:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except:
            return date_str

    def get_tp_imp_desc(self, tp_imp):
        """获取打印类型描述"""
        types = {
            "1": "普通打印",
            "2": "DANFE简化",
            "3": "DANFE NFC-e",
            "4": "DANFE NFC-e手机",
            "5": "DANFE NFC-e邮件"
        }
        return types.get(tp_imp, "未知")
        
    def get_tp_emis_desc(self, tp_emis):
        """获取出具类型描述"""
        types = {
            "1": "正常出具",
            "2": "应急出具",
            "3": "SCAN出具",
            "4": "DPEC出具",
            "5": "应急出具FS-DA",
            "6": "SVC-AN出具",
            "7": "SVC-RS出具",
            "8": "SVCSP出具",
            "9": "离线应急"
        }
        return types.get(tp_emis, "未知")

    def export_to_excel(self):
        """导出数据到Excel文件"""
        if not self.invoice_data:
            messagebox.showwarning("警告", "没有数据可以导出")
            return
        
        try:
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            
            # 创建发票明细表
            self.create_invoice_detail_sheet(wb)
            
            # 创建发票汇总表
            self.create_invoice_summary_sheet(wb)
            
            # 删除默认的Sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # 保存文件
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="保存Excel文件"
            )
            
            if filename:
                wb.save(filename)
                messagebox.showinfo("成功", "数据已成功导出到Excel文件")
        
        except Exception as e:
            messagebox.showerror("错误", f"导出Excel时出错: {str(e)}")

    def create_invoice_detail_sheet(self, wb):
        """创建发票明细表"""
        ws = wb.create_sheet("发票明细")
        
        # 设置列标题
        headers = [
            "发票号码", "发票日期", "客户名称", "客户CNPJ", "客户地址",
            "产品代码", "产品名称", "NCM", "数量", "单价", "总价",
            "ICMS税率", "ICMS税额", "PIS税率", "PIS税额",
            "COFINS税率", "COFINS税额"
        ]
        
        # 设置标题样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入标题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 写入数据
        row = 2
        for item in self.tree.get_children():
            values = self.tree.item(item)["values"]
            for col, value in enumerate(values[:len(headers)], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left")
            row += 1

    def create_invoice_summary_sheet(self, wb):
        """创建发票汇总表"""
        ws = wb.create_sheet("发票汇总")
        
        # 设置列标题
        headers = [
            "发票号码", "发票日期", "客户名称", "客户CNPJ", "客户地址",
            "产品总价", "运费", "折扣", "ICMS税率", "ICMS税额",
            "PIS税率", "PIS税额", "COFINS税率", "COFINS税额"
        ]
        
        # 设置标题样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入标题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 写入数据
        row = 2
        for item in self.invoice_tree.get_children():
            values = self.invoice_tree.item(item)["values"]
            for col, value in enumerate(values[:len(headers)], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left")
            row += 1

    def _on_mousewheel(self, event):
        """处理鼠标滚轮事件"""
        # 获取当前操作系统
        if sys.platform.startswith('win'):
            # Windows系统
            delta = -int(event.delta/120)
        else:
            # Linux和macOS系统
            delta = event.delta
        
        # 计算滚动的单位（每次滚动3行）
        scroll_unit = delta * 3
        
        # 根据事件的widget来确定要滚动的组件
        widget = event.widget
        
        if widget == self.tree:
            # 数据表格滚动
            self.tree.yview_scroll(scroll_unit, "units")
        elif widget == self.invoice_tree:
            # 发票表格滚动
            self.invoice_tree.yview_scroll(scroll_unit, "units")
        elif widget == self.file_tree:
            # 文件列表滚动
            self.file_tree.yview_scroll(scroll_unit, "units")
            
    def _bind_mousewheel(self, event):
        """绑定鼠标滚轮事件"""
        # 当鼠标进入组件时绑定滚轮事件
        event.widget.bind_all("<MouseWheel>", self._on_mousewheel)
        
    def _unbind_mousewheel(self, event):
        """解绑鼠标滚轮事件"""
        # 当鼠标离开组件时解绑滚轮事件
        event.widget.unbind_all("<MouseWheel>")

    def _on_shift_mousewheel(self, event):
        """处理Shift+鼠标滚轮事件（用于水平滚动）"""
        # 获取当前操作系统
        if sys.platform.startswith('win'):
            # Windows系统
            delta = -int(event.delta/120)
        else:
            # Linux和macOS系统
            delta = event.delta
        
        # 计算滚动的单位（每次滚动3列）
        scroll_unit = delta * 3
        
        # 根据事件的widget来确定要滚动的组件
        widget = event.widget
        
        if widget == self.tree:
            # 数据表格水平滚动
            self.tree.xview_scroll(scroll_unit, "units")
        elif widget == self.invoice_tree:
            # 发票表格水平滚动
            self.invoice_tree.xview_scroll(scroll_unit, "units")
            
    def _bind_shift_mousewheel(self, event):
        """绑定Shift+鼠标滚轮事件"""
        # 当鼠标进入组件时绑定Shift+滚轮事件
        event.widget.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)
        
    def _unbind_shift_mousewheel(self, event):
        """解绑Shift+鼠标滚轮事件"""
        # 当鼠标离开组件时解绑Shift+滚轮事件
        event.widget.unbind_all("<Shift-MouseWheel>")

    def _debounced_scroll(self, scrollbar, *args):
        """防抖动滚动处理"""
        if not hasattr(self, '_scroll_after_id'):
            self._scroll_after_id = None
            
        # 取消之前的待执行滚动
        if self._scroll_after_id:
            self.after_cancel(self._scroll_after_id)
            
        # 延迟执行滚动
        def update_scroll():
            scrollbar.set(*args)
            self._scroll_after_id = None
            
        self._scroll_after_id = self.after(10, update_scroll)

    def process_xml(self, tree):
        """处理XML文件"""
        try:
            # 定义命名空间
            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
            
            # 获取根元素和必要节点
            root = tree.getroot()
            nfe = root.find('.//nfe:NFe', ns)
            if nfe is None:
                raise ValueError("找不到NFe节点")
            
            inf_nfe = nfe.find('.//nfe:infNFe', ns)
            if inf_nfe is None:
                raise ValueError("找不到infNFe节点")
            
            # 获取ide节点（标识信息）
            ide = inf_nfe.find('nfe:ide', ns)
            if ide is None:
                raise ValueError("找不到ide节点")
            
            # 获取基本信息
            invoice_number = self.safe_find_text(ide, 'nfe:nNF', ns)
            series = self.safe_find_text(ide, 'nfe:serie', ns)
            date = self.format_date(self.safe_find_text(ide, 'nfe:dhEmi', ns))
            nat_op = self.safe_find_text(ide, 'nfe:natOp', ns)
            tp_imp = self.get_tp_imp_desc(self.safe_find_text(ide, 'nfe:tpImp', ns))
            tp_emis = self.get_tp_emis_desc(self.safe_find_text(ide, 'nfe:tpEmis', ns))
            
            # 获取客户信息
            dest = inf_nfe.find('nfe:dest', ns)
            if dest is None:
                raise ValueError("找不到dest节点")
            
            customer_name = self.safe_find_text(dest, 'nfe:xNome', ns)
            customer_cnpj = self.safe_find_text(dest, 'nfe:CNPJ', ns)
            
            # 获取客户地址
            ender_dest = dest.find('nfe:enderDest', ns)
            if ender_dest is not None:
                address_parts = [
                    self.safe_find_text(ender_dest, 'nfe:xLgr', ns),
                    self.safe_find_text(ender_dest, 'nfe:nro', ns),
                    self.safe_find_text(ender_dest, 'nfe:xBairro', ns),
                    self.safe_find_text(ender_dest, 'nfe:xMun', ns),
                    self.safe_find_text(ender_dest, 'nfe:UF', ns),
                    self.safe_find_text(ender_dest, 'nfe:CEP', ns)
                ]
                customer_address = ', '.join(filter(None, address_parts))
            else:
                customer_address = ""
            
            # 获取总计信息
            total = inf_nfe.find('.//nfe:total/nfe:ICMSTot', ns)
            if total is None:
                raise ValueError("找不到total节点")
            
            # 获取商品信息
            items = inf_nfe.findall('.//nfe:det', ns)
            
            # 更新基本统计
            print(f"处理发票 {invoice_number}")
            print(f"更新前统计数据: {self.summary_data}")
            
            self.summary_data["total_invoices"] += 1
            self.summary_data["total_products"] += len(items)
            
            # 更新金额统计
            total_amount = self.safe_find_decimal(total, 'nfe:vNF', ns)
            total_icms = self.safe_find_decimal(total, 'nfe:vICMS', ns)
            total_pis = self.safe_find_decimal(total, 'nfe:vPIS', ns)
            total_cofins = self.safe_find_decimal(total, 'nfe:vCOFINS', ns)
            total_discount = self.safe_find_decimal(total, 'nfe:vDesc', ns)
            
            self.summary_data["total_amount"] += total_amount
            self.summary_data["total_icms"] += total_icms
            self.summary_data["total_pis"] += total_pis
            self.summary_data["total_cofins"] += total_cofins
            self.summary_data["total_tax"] += (total_icms + total_pis + total_cofins)
            self.summary_data["total_discount"] += total_discount
            
            # 更新客户统计
            if customer_cnpj:
                self.summary_data["unique_customers"].add(customer_cnpj)
            
            # 更新发票数据
            self.invoice_data[invoice_number] = {
                "invoice_date": date,
                "customer_name": customer_name,
                "customer_cnpj": customer_cnpj,
                "customer_address": customer_address,
                "total_amount": total_amount,
                "freight": self.safe_find_decimal(total, 'nfe:vFrete', ns),
                "discount": total_discount,
                "icms_rate": "0",
                "icms_amount": total_icms,
                "pis_rate": "0",
                "pis_amount": total_pis,
                "cofins_rate": "0",
                "cofins_amount": total_cofins
            }
            
            # 处理商品信息
            for item in items:
                prod = item.find('nfe:prod', ns)
                if prod is not None:
                    # 获取商品基本信息
                    product_code = self.safe_find_text(prod, 'nfe:cProd', ns)
                    product_name = self.safe_find_text(prod, 'nfe:xProd', ns)
                    ncm = self.safe_find_text(prod, 'nfe:NCM', ns)
                    quantity = self.safe_find_decimal(prod, 'nfe:qCom', ns)
                    unit_price = self.safe_find_decimal(prod, 'nfe:vUnCom', ns)
                    total_price = self.safe_find_decimal(prod, 'nfe:vProd', ns)
                    
                    # 获取ICMS信息
                    icms = item.find('.//nfe:ICMS', ns)
                    icms_rate = "0"
                    icms_amount = Decimal('0')
                    
                    if icms is not None:
                        for icms_type in ['ICMS00', 'ICMS10', 'ICMS20', 'ICMS30', 'ICMS40', 'ICMS50', 'ICMS60', 'ICMS70', 'ICMS90']:
                            icms_group = icms.find(f'nfe:{icms_type}', ns)
                            if icms_group is not None:
                                icms_rate = self.safe_find_text(icms_group, 'nfe:pICMS', ns)
                                icms_amount = self.safe_find_decimal(icms_group, 'nfe:vICMS', ns)
                                break
                    
                    # 获取PIS信息
                    pis = item.find('.//nfe:PIS', ns)
                    pis_rate = "0"
                    pis_amount = Decimal('0')
                    
                    if pis is not None:
                        pis_group = pis.find('nfe:PISAliq', ns)
                        if pis_group is not None:
                            pis_rate = self.safe_find_text(pis_group, 'nfe:pPIS', ns)
                            pis_amount = self.safe_find_decimal(pis_group, 'nfe:vPIS', ns)
                    
                    # 获取COFINS信息
                    cofins = item.find('.//nfe:COFINS', ns)
                    cofins_rate = "0"
                    cofins_amount = Decimal('0')
                    
                    if cofins is not None:
                        cofins_group = cofins.find('nfe:COFINSAliq', ns)
                        if cofins_group is not None:
                            cofins_rate = self.safe_find_text(cofins_group, 'nfe:pCOFINS', ns)
                            cofins_amount = self.safe_find_decimal(cofins_group, 'nfe:vCOFINS', ns)
                    
                    # 添加到数据表格
                    self.tree.insert("", tk.END, values=(
                        invoice_number, series, date, nat_op, tp_imp, tp_emis,
                        customer_name, customer_cnpj, customer_address,
                        product_code, product_name, ncm,
                        float(quantity), float(unit_price), float(total_price),
                        icms_rate, float(icms_amount),
                        pis_rate, float(pis_amount),
                        cofins_rate, float(cofins_amount)
                    ))
                    
                    # 更新商品统计
                    if product_code:
                        self.summary_data["unique_products"].add(product_code)
                    
                    # 更新价格统计
                    if unit_price > Decimal('0'):
                        if self.summary_data["max_product_price"] == Decimal('0') or unit_price > self.summary_data["max_product_price"]:
                            self.summary_data["max_product_price"] = unit_price
                        if self.summary_data["min_product_price"] == Decimal('0') or unit_price < self.summary_data["min_product_price"]:
                            self.summary_data["min_product_price"] = unit_price
            
            # 计算平均价格
            if self.summary_data["total_products"] > 0:
                self.summary_data["avg_product_price"] = self.summary_data["total_amount"] / Decimal(str(self.summary_data["total_products"]))
            
            # 计算税收占比
            if self.summary_data["total_amount"] > Decimal('0'):
                self.summary_data["tax_percentage"] = (self.summary_data["total_tax"] / self.summary_data["total_amount"]) * Decimal('100')
            
            print(f"更新后统计数据: {self.summary_data}")
            
            # 更新发票列表
            self.invoice_tree.insert("", tk.END, values=(
                invoice_number,
                date,
                customer_name,
                customer_cnpj,
                customer_address,
                f"R$ {float(total_amount):,.2f}",
                f"R$ {float(self.invoice_data[invoice_number]['freight']):,.2f}",
                f"R$ {float(total_discount):,.2f}",
                f"{icms_rate}%",
                f"R$ {float(total_icms):,.2f}",
                f"{pis_rate}%",
                f"R$ {float(total_pis):,.2f}",
                f"{cofins_rate}%",
                f"R$ {float(total_cofins):,.2f}"
            ))
            
            # 更新统计显示
            self.update_summary_display()
            
            # 强制更新界面
            self.update()
            self.update_idletasks()
            
        except Exception as e:
            print(f"处理XML时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def update_invoice_display(self):
        """更新发票显示"""
        try:
            # 清除现有数据
            for item in self.invoice_tree.get_children():
                self.invoice_tree.delete(item)
            
            # 重新填充数据
            for invoice_number, data in self.invoice_data.items():
                self.invoice_tree.insert("", tk.END, values=(
                    invoice_number,
                    data["invoice_date"],
                    data["customer_name"],
                    data["customer_cnpj"],
                    data["customer_address"],
                    f"R$ {float(data['total_amount']):,.2f}",
                    f"R$ {float(data['freight']):,.2f}",
                    f"R$ {float(data['discount']):,.2f}",
                    f"{data['icms_rate']}%",
                    f"R$ {float(data['icms_amount']):,.2f}",
                    f"{data['pis_rate']}%",
                    f"R$ {float(data['pis_amount']):,.2f}",
                    f"{data['cofins_rate']}%",
                    f"R$ {float(data['cofins_amount']):,.2f}"
                ))
        except Exception as e:
            print(f"更新发票显示时出错: {str(e)}")

    def create_summary_page(self):
        """创建统计分析页面"""
        summary_page = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(summary_page, text="统计分析")
        
        # 创建左右布局
        left_frame = ttk.Frame(summary_page, style="Card.TFrame")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 5), pady=10)
        
        right_frame = ttk.Frame(summary_page, style="Card.TFrame")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 10), pady=10)
        
        # 创建统计卡片
        self.create_basic_stats_card(left_frame)
        self.create_tax_stats_card(left_frame)
        self.create_product_stats_card(right_frame)
        self.create_customer_stats_card(right_frame)

    def create_basic_stats_card(self, parent):
        """创建基本统计卡片"""
        card = ttk.LabelFrame(parent, text="基本统计", style="Card.TLabelframe")
        card.pack(fill=tk.X, pady=(0, 10), ipadx=10, ipady=5)
        
        stats = [
            ("total_invoices", "总发票数", "件"),
            ("total_products", "总商品数", "件"),
            ("total_amount", "总金额", "R$", True)
        ]
        
        for key, label, unit, *args in stats:
            self.create_stat_row(card, key, label, unit, is_currency=bool(args))

    def create_tax_stats_card(self, parent):
        """创建税收统计卡片"""
        card = ttk.LabelFrame(parent, text="税收统计", style="Card.TLabelframe")
        card.pack(fill=tk.X, pady=(0, 10), ipadx=10, ipady=5)
        
        stats = [
            ("total_icms", "ICMS税额", "R$", True),
            ("total_pis", "PIS税额", "R$", True),
            ("total_cofins", "COFINS税额", "R$", True),
            ("total_tax", "总税额", "R$", True),
            ("tax_percentage", "税收占比", "%")
        ]
        
        for key, label, unit, *args in stats:
            self.create_stat_row(card, key, label, unit, is_currency=bool(args))

    def create_product_stats_card(self, parent):
        """创建商品统计卡片"""
        card = ttk.LabelFrame(parent, text="商品统计", style="Card.TLabelframe")
        card.pack(fill=tk.X, pady=(0, 10), ipadx=10, ipady=5)
        
        stats = [
            ("unique_products", "商品种类", "种"),
            ("avg_product_price", "平均单价", "R$", True),
            ("max_product_price", "最高单价", "R$", True),
            ("min_product_price", "最低单价", "R$", True),
            ("total_discount", "总折扣", "R$", True)
        ]
        
        for key, label, unit, *args in stats:
            self.create_stat_row(card, key, label, unit, is_currency=bool(args))

    def create_customer_stats_card(self, parent):
        """创建客户统计卡片"""
        card = ttk.LabelFrame(parent, text="客户统计", style="Card.TLabelframe")
        card.pack(fill=tk.X, pady=(0, 10), ipadx=10, ipady=5)
        
        self.create_stat_row(card, "unique_customers", "客户数量", "个")

    def create_stat_row(self, parent, key, label_text, unit, is_currency=False):
        """创建统计行"""
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.pack(fill=tk.X, pady=2)
        
        # 创建标签容器
        label_container = ttk.Frame(frame, style="Card.TFrame")
        label_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # 添加图标
        icon_label = ttk.Label(
            label_container,
            text="📊",
            style="Header.TLabel"
        )
        icon_label.pack(side=tk.LEFT, padx=(0, 5))
        
        # 添加标签文本
        label = ttk.Label(
            label_container,
            text=f"{label_text}:",
            style="Header.TLabel"
        )
        label.pack(side=tk.LEFT)
        
        # 创建值容器
        value_container = ttk.Frame(frame, style="Card.TFrame")
        value_container.pack(side=tk.RIGHT, padx=5)
        
        # 创建值标签
        value_label = ttk.Label(
            value_container,
            text="0" if not is_currency else "R$ 0.00",
            style="Value.TLabel"
        )
        value_label.pack(side=tk.LEFT)
        
        # 添加单位标签
        if unit:
            unit_label = ttk.Label(
                value_container,
                text=f" {unit}",
                style="Value.TLabel"
            )
            unit_label.pack(side=tk.LEFT)
        
        # 存储标签引用
        self.summary_labels[key] = value_label
        
        return value_label

    def update_summary_display(self):
        """更新统计显示"""
        try:
            print("开始更新统计显示")
            print(f"当前统计数据: {self.summary_data}")
            
            # 更新所有统计标签
            for key, value in self.summary_data.items():
                if key in self.summary_labels:
                    label = self.summary_labels[key]
                    if isinstance(value, set):
                        label.configure(text=str(len(value)))
                    elif isinstance(value, Decimal):
                        if key == "tax_percentage":
                            label.configure(text=f"{float(value):.1f}")
                        else:
                            label.configure(text=f"R$ {float(value):,.2f}")
                    else:
                        label.configure(text=str(value))
                    print(f"更新 {key}: {value}")
            
            # 强制更新显示
            self.update_idletasks()
            
        except Exception as e:
            print(f"更新统计显示时出错: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    app = NFeAnalyzer()
    app.mainloop() 
